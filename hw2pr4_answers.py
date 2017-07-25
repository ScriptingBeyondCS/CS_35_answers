from openpyxl import load_workbook

def decodeText():

    text = 'rlrrrrwrefrxrrjrrnmhrretyrrwmrroorrrrjwlrretyrlrrrrrjwlrooiirbrbbrrjwlrurnrretrtrooiirrrreryrrerrwmrrrjwlrwrefrtreertreeryrrerwrefrrwmrrrjwlrwrorrrjwlrretyrurnrrrjwlrrnmhrhrrrryrrerrjwlroorrrxrrjrrwmrryrreryrrerurnrrrjwlrwrorrrjwlrlkrkrlrrrrrjwlrlkrkrrjwlrwrefrooiirrjwlrrnmhrretyrrjwlrprrrrwrefrurnrrlkrkroorrrhrrrrrjwlrwrorrrjwlrrnmhrhrrrryrrerurnrrrjwlrrnmhretrtrrwmrrurnrrrjwlrrnmhrhrrrryrrerooiirrjwlrnenerlkrkrirrrrhrrrrrnmhrrjwlrirrrrrwmrryrreryrrerurnr'

    # factors of my number 
    # appear on the screen 
    # if i am to vanish 
    # then turn them light green


    wb = load_workbook('message.xlsx')
    ws = wb.active

    dictionary = {}
    for row in ws.iter_rows():
        cell0 = row[0].value
        cell1 = row[1].value
        dictionary[cell0] = cell1

    new_text = ''
    for i in range(0, len(text), 5):
        x = text[i:i+5]
        if x.lower() in dictionary:
            new_text += dictionary[x.lower()]
        else:
            new_text += x

    # TO ENCODE

    # new_text = ''
    # for x in text:
        
    #     if x.lower() in dictionary:
    #         new_text += dictionary[x.lower()]
    #     else:
    #         new_text += x

    print(new_text)
    
    
def decodeNumbers():

    f = open('message.txt', 'r')
    text = f.read()
    f.close()

    html = ''

    for x in text:
        if x in '12367':
            html += '<span style="color:{0};">{1}</span>'.format("lightgreen", x)
        elif x == '\n':
            html += "<br>"
        else:
            html += x


    f = open('message.html', 'w')
    f.write(html)
    f.close()


